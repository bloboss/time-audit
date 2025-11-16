"""Excel export functionality with charts and formatting."""

from collections import defaultdict
from datetime import datetime
from typing import Any, Optional

from time_audit.core.models import Entry
from time_audit.export_import.base import Exporter


class ExcelExporter(Exporter):
    """Export time tracking data to Excel format with charts."""

    def get_file_extension(self) -> str:
        """Get Excel file extension.

        Returns:
            '.xlsx'
        """
        return ".xlsx"

    def export_entries(
        self,
        entries: list[Entry],
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        **kwargs: Any,
    ) -> None:
        """Export entries to Excel file with formatting and charts.

        Args:
            entries: List of entries to export
            start_date: Optional start date filter
            end_date: Optional end date filter
            **kwargs: Additional options
                - include_charts (bool): Include charts (default: True)
                - include_summary (bool): Include summary sheet (default: True)
        """
        self.ensure_output_path()

        # Filter entries
        filtered_entries = self.filter_entries(entries, start_date, end_date)

        # Try to import openpyxl
        try:
            import openpyxl  # type: ignore[import-untyped]
            from openpyxl.chart import (  # type: ignore[import-untyped]  # noqa: F401
                BarChart,
                PieChart,
                Reference,
            )
            from openpyxl.styles import (  # type: ignore[import-untyped]  # noqa: F401
                Alignment,
                Font,
                PatternFill,
            )
            from openpyxl.utils import (
                get_column_letter,  # type: ignore[import-untyped]  # noqa: F401
            )
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. " "Install with: pip install openpyxl"
            )

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create entries sheet
        self._create_entries_sheet(wb, filtered_entries)

        # Create summary sheet if requested
        if kwargs.get("include_summary", True):
            self._create_summary_sheet(wb, filtered_entries, kwargs.get("include_charts", True))

        # Save workbook
        wb.save(self.output_path)

    def _create_entries_sheet(self, wb: Any, entries: list[Entry]) -> None:
        """Create detailed entries sheet.

        Args:
            wb: Workbook object
            entries: List of entries
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        ws = wb.create_sheet("Entries")

        # Define headers
        headers = [
            "Task",
            "Start Time",
            "End Time",
            "Duration (hrs)",
            "Active (hrs)",
            "Idle (%)",
            "Project",
            "Category",
            "Tags",
            "Notes",
        ]

        # Write headers with formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write entry data
        for row, entry in enumerate(entries, start=2):
            ws.cell(row, 1, entry.task_name)
            ws.cell(row, 2, entry.start_time.strftime("%Y-%m-%d %H:%M:%S"))

            if entry.end_time:
                ws.cell(row, 3, entry.end_time.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                ws.cell(row, 3, "Running")

            if entry.duration_seconds:
                ws.cell(row, 4, round(entry.duration_seconds / 3600, 2))
            else:
                ws.cell(row, 4, "-")

            if entry.active_duration_seconds:
                ws.cell(row, 5, round(entry.active_duration_seconds / 3600, 2))
            else:
                ws.cell(row, 5, "-")

            if entry.idle_percentage is not None:
                ws.cell(row, 6, round(entry.idle_percentage, 1))
            else:
                ws.cell(row, 6, "-")

            ws.cell(row, 7, entry.project or "")
            ws.cell(row, 8, entry.category or "")
            ws.cell(row, 9, ", ".join(entry.tags) if entry.tags else "")
            ws.cell(row, 10, entry.notes or "")

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    def _create_summary_sheet(
        self, wb: Any, entries: list[Entry], include_charts: bool = True
    ) -> None:
        """Create summary sheet with aggregated data and charts.

        Args:
            wb: Workbook object
            entries: List of entries
            include_charts: Whether to include charts
        """
        from openpyxl.chart import PieChart, Reference
        from openpyxl.styles import Font

        ws = wb.create_sheet("Summary", 0)  # Insert as first sheet

        # Calculate summary statistics
        total_duration = sum(e.duration_seconds for e in entries if e.duration_seconds)
        total_active = sum(e.active_duration_seconds for e in entries if e.active_duration_seconds)

        # Time by task
        task_time: dict[str, float] = defaultdict(float)
        for entry in entries:
            if entry.duration_seconds:
                task_time[entry.task_name] += entry.duration_seconds / 3600

        # Time by project
        project_time: dict[str, float] = defaultdict(float)
        for entry in entries:
            if entry.duration_seconds:
                project = entry.project or "No Project"
                project_time[project] += entry.duration_seconds / 3600

        # Time by category
        category_time: dict[str, float] = defaultdict(float)
        for entry in entries:
            if entry.duration_seconds:
                category = entry.category or "Uncategorized"
                category_time[category] += entry.duration_seconds / 3600

        # Write overall statistics
        header_font = Font(bold=True, size=14)
        ws["A1"] = "Time Tracking Summary"
        ws["A1"].font = header_font

        ws["A3"] = "Total Time Tracked:"
        ws["B3"] = f"{total_duration / 3600:.2f} hours"

        ws["A4"] = "Total Active Time:"
        ws["B4"] = f"{total_active / 3600:.2f} hours"

        ws["A5"] = "Total Entries:"
        ws["B5"] = len(entries)

        # Time by Task
        row = 7
        ws[f"A{row}"] = "Time by Task"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        ws[f"A{row}"] = "Task"
        ws[f"B{row}"] = "Hours"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)
        row += 1

        for task, hours in sorted(task_time.items(), key=lambda x: x[1], reverse=True):
            ws[f"A{row}"] = task
            ws[f"B{row}"] = round(hours, 2)
            row += 1

        # Time by Project
        row += 1
        ws[f"A{row}"] = "Time by Project"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        ws[f"A{row}"] = "Project"
        ws[f"B{row}"] = "Hours"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)
        row += 1

        project_start_row = row
        for project, hours in sorted(project_time.items(), key=lambda x: x[1], reverse=True):
            ws[f"A{row}"] = project
            ws[f"B{row}"] = round(hours, 2)
            row += 1

        # Add charts if requested and data available
        if include_charts and task_time:
            # Project pie chart
            chart = PieChart()
            chart.title = "Time by Project"
            data = Reference(ws, min_col=2, min_row=project_start_row - 1, max_row=row - 1)
            labels = Reference(ws, min_col=1, min_row=project_start_row, max_row=row - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.height = 10
            chart.width = 15
            ws.add_chart(chart, "D3")

        # Auto-size columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
